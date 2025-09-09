"""
Document Parser using Gemini Vision API with File Upload
Extracts tables from PDF documents using Google's Gemini 2.5 Pro model
Automatically detects table names and handles multi-page tables
"""

import google.generativeai as genai
import pandas as pd
import json
import os
import time  # For retries
import pymupdf as fitz  # PyMuPDF for fast PDF processing
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure API from environment variables
API_KEY = os.getenv('GEMINI_API_KEY')
if not API_KEY:
    raise ValueError("GEMINI_API_KEY not found in environment variables. Please check your .env file.")

genai.configure(api_key=API_KEY)
model = genai.GenerativeModel('gemini-2.5-pro')  # Use 2.5 Pro for better accuracy

def parse_pdf_with_gemini(pdf_path, output_excel=None, retries=2):
    """
    Process PDF page by page using PyMuPDF + Gemini for fast extraction.
    Saves results to Excel in real-time as tables are found.
    Returns dict of DataFrames.
    """
    print(f"Processing PDF: {pdf_path}")
    
    # Open PDF with PyMuPDF
    doc = fitz.open(pdf_path)
    all_tables = {}
    
    prompt = """
Analyze this page of a financial statement. Extract any tables present as JSON with proper sub-heading structure:

{
  "table_name": "Main table name (e.g., 'Fees', 'Interchange Program Fees')",
  "headers": ["Column1", "Column2", ...],
  "rows": [
    ["SUBHEADING", "", "", ...],  // Single word sub-heading with empty cells
    ["data1", "data2", "data3", ...],  // Data rows under sub-heading
    ["data4", "data5", "data6", ...],
    ["SUBHEADING2", "", "", ...],  // Another sub-heading
    ["data7", "data8", "data9", ...]
  ]
}

IMPORTANT RULES:
1. If you see single words like "VISA", "MASTERCARD", "DISCOVER", "AMEX", "ACQ" in a row with empty cells after, treat them as SUBHEADINGS
2. Keep sub-headings as separate rows with empty cells in other columns
3. Group related data under their sub-headings
4. For fees tables, use "Fees" as main table name, not "Transaction Fees" or "Interchange Fees"
5. For interchange program fees, use "Interchange Program Fees" as main table name
6. Preserve the hierarchical structure with sub-headings
7. ALWAYS ensure headers array has the SAME LENGTH as each row array
8. If a table continues from previous page, use the EXACT SAME table name and headers
9. Fill empty cells with empty strings "" not null values

If no tables found, return: {"table_name": null, "rows": []}
Focus on accuracy for numbers and dates.
"""
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        print(f"Processing page {page_num + 1} of {len(doc)}")
        
        # Convert page to image using PyMuPDF (fast)
        mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better quality
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes("png")
        
        # Create image part for Gemini
        img_part = {'mime_type': 'image/png', 'data': img_data}
        
        for attempt in range(retries):
            try:
                response = model.generate_content([prompt, img_part])
                
                # Clean response
                json_str = response.text.strip()
                if '```json' in json_str:
                    json_str = json_str.split('```json')[1].split('```')[0]
                elif '```' in json_str:
                    json_str = json_str.split('```')[1].split('```')[0]
                
                data = json.loads(json_str)
                
                # Handle both single object and array responses
                if isinstance(data, list):
                    tables_found = False
                    for item in data:
                        if item.get('table_name') and item.get('rows'):
                            table_name = item['table_name']
                            df = pd.DataFrame(item['rows'], columns=item['headers'])
                            
                            # Check if this is a continuation of an existing table
                            if table_name in all_tables:
                                # Merge with existing table (continuation)
                                existing_df = all_tables[table_name]
                                # Check if headers match (for table continuation)
                                if list(df.columns) == list(existing_df.columns):
                                    try:
                                        # Validate DataFrames before merging
                                        if len(df) > 0 and len(existing_df) > 0:
                                            # Merge rows, preserving sub-headings
                                            all_tables[table_name] = pd.concat([existing_df, df], ignore_index=True)
                                            print(f"✓ Merged continuation into table: {table_name}")
                                        else:
                                            print(f"⚠ Skipping merge for {table_name} - empty DataFrame")
                                    except Exception as e:
                                        print(f"⚠ Error merging {table_name}: {e}")
                                        # Fallback: create new table with page number
                                        table_name = f"{table_name}_Page{page_num + 1}"
                                        all_tables[table_name] = df
                                        print(f"✓ Created new table: {table_name}")
                                else:
                                    # Different structure, create new table with page number
                                    table_name = f"{table_name}_Page{page_num + 1}"
                                    all_tables[table_name] = df
                                    print(f"✓ Created new table: {table_name}")
                            else:
                                # New table
                                all_tables[table_name] = df
                                print(f"✓ Extracted table: {table_name}")
                            
                            tables_found = True
                            
                            # Save to Excel in real-time
                            if output_excel:
                                save_tables_to_excel(all_tables, output_excel)
                    
                    if not tables_found:
                        print(f"✗ No tables found on page {page_num + 1}")
                        
                elif isinstance(data, dict) and data.get('table_name') and data.get('rows'):
                    table_name = data['table_name']
                    df = pd.DataFrame(data['rows'], columns=data['headers'])
                    
                    # Check if this is a continuation of an existing table
                    if table_name in all_tables:
                        # Merge with existing table (continuation)
                        existing_df = all_tables[table_name]
                        # Check if headers match (for table continuation)
                        if list(df.columns) == list(existing_df.columns):
                            try:
                                # Validate DataFrames before merging
                                if len(df) > 0 and len(existing_df) > 0:
                                    # Merge rows
                                    all_tables[table_name] = pd.concat([existing_df, df], ignore_index=True)
                                    print(f"✓ Merged continuation into table: {table_name}")
                                else:
                                    print(f"⚠ Skipping merge for {table_name} - empty DataFrame")
                            except Exception as e:
                                print(f"⚠ Error merging {table_name}: {e}")
                                # Fallback: create new table with page number
                                table_name = f"{table_name}_Page{page_num + 1}"
                                all_tables[table_name] = df
                                print(f"✓ Created new table: {table_name}")
                        else:
                            # Different structure, create new table with page number
                            table_name = f"{table_name}_Page{page_num + 1}"
                            all_tables[table_name] = df
                            print(f"✓ Created new table: {table_name}")
                    else:
                        # New table
                        all_tables[table_name] = df
                        print(f"✓ Extracted table: {table_name}")
                    
                    # Save to Excel in real-time
                    if output_excel:
                        save_tables_to_excel(all_tables, output_excel)
                else:
                    print(f"✗ No tables found on page {page_num + 1}")
                
                break  # Success, no need to retry
                
            except json.JSONDecodeError as e:
                print(f"JSON parsing error on page {page_num + 1}, attempt {attempt + 1}: {e}")
                if attempt < retries - 1:
                    time.sleep(1)
            except Exception as e:
                print(f"Error on page {page_num + 1}, attempt {attempt + 1}: {e}")
                if attempt < retries - 1:
                    time.sleep(1)
    
    doc.close()
    
    # Clean and validate data
    for name, df in all_tables.items():
        for col in df.columns:
            if any(keyword in col.lower() for keyword in ['amount', 'fee', 'total', 'count', 'number']):
                df[col] = pd.to_numeric(df[col], errors='coerce')
    
    print(f"Successfully extracted {len(all_tables)} tables")
    return all_tables

def save_tables_to_excel(tables, output_excel):
    """
    Save tables to Excel file with sub-heading formatting (helper function for real-time saving)
    """
    try:
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            for sheet, df in tables.items():
                # Clean sheet name for Excel compatibility
                clean_name = "".join(c for c in sheet if c.isalnum() or c in (' ', '-', '_')).strip()[:31]
                df.to_excel(writer, sheet_name=clean_name, index=False)
                
                # Format sub-headings (single word in first column with empty cells)
                worksheet = writer.sheets[clean_name]
                for row_idx, row in df.iterrows():
                    first_cell = str(row.iloc[0]) if len(row) > 0 else ""
                    # Check if this looks like a sub-heading (single word, rest empty)
                    if (first_cell and 
                        first_cell.upper() in ['VISA', 'MASTERCARD', 'DISCOVER', 'AMEX', 'ACQ', 'DEBIT', 'CREDIT'] and
                        all(str(cell).strip() == '' or str(cell) == 'nan' for cell in row.iloc[1:])):
                        # Format as sub-heading (bold, background color)
                        from openpyxl.styles import Font, PatternFill
                        for col_idx in range(1, len(row) + 1):
                            cell = worksheet.cell(row=row_idx + 2, column=col_idx)  # +2 because Excel is 1-indexed and has header
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        print(f"💾 Saved {len(tables)} tables to Excel with sub-heading formatting")
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def automate_to_excel(pdf_path, output_excel):
    """
    Process a single PDF and export tables to Excel
    """
    try:
        # Pass output_excel to enable real-time saving
        tables = parse_pdf_with_gemini(pdf_path, output_excel)
        
        if not tables:
            print(f"No tables found in {pdf_path}")
            return
        
        print(f"✓ Final export: {len(tables)} tables saved to {output_excel}")
        
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")

def process_folder(input_folder, output_folder):
    """
    Process all PDFs in a folder and export to Excel files
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Created output directory: {output_folder}")
    
    pdf_files = [f for f in os.listdir(input_folder) if f.endswith('.pdf')]
    
    if not pdf_files:
        print(f"No PDF files found in {input_folder}")
        return
    
    print(f"Found {len(pdf_files)} PDF files to process")
    
    for filename in pdf_files:
            pdf_path = os.path.join(input_folder, filename)
        excel_path = os.path.join(output_folder, f"{os.path.splitext(filename)[0]}.xlsx")
        
        print(f"\n--- Processing {filename} ---")
        automate_to_excel(pdf_path, excel_path)

def process_single_pdf(pdf_path, output_excel=None):
    """
    Process a single PDF file
    """
    # If just filename is provided, look in input_pdfs folder
    if not os.path.dirname(pdf_path):
        pdf_path = os.path.join('input_pdfs', pdf_path)
    
    if not os.path.exists(pdf_path):
        print(f"PDF file not found: {pdf_path}")
        return
    
    if output_excel is None:
        # Create output filename based on input
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_excel = os.path.join('output_excels', f"{base_name}_tables.xlsx")
    
    print(f"Processing single PDF: {pdf_path}")
    automate_to_excel(pdf_path, output_excel)

def main():
    """
    Main function to run the document parser
    """
    print("=== Document Parser using Gemini Vision API ===")
    print("Make sure to set your API_KEY in the script!")
    
    # Check if API key is set
    if API_KEY == "your_api_key_here":
        print("❌ Please set your Gemini API key in the script!")
        print("Get your API key from: https://makersuite.google.com/app/apikey")
        return
    
    # Process all PDFs in the input folder
    process_folder('input_pdfs/', 'output_excels/')
    print("\n=== Processing Complete ===")

if __name__ == '__main__':
    # Process a single PDF file
    process_single_pdf('4445065529422-BIMERBIL-09-30-2024.pdf')
    
    # To process all PDFs in a folder, uncomment the line below:
    # main()